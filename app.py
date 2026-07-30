import streamlit as st
import yfinance as ticker_data
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from screening_logic import screen_company

# --- PAGE CONFIG (neutral, no internal version numbers) ---
st.set_page_config(
    page_title="HalalQuant | Independent Financial Screening",
    page_icon="📊",
    layout="wide"
)

# --- CUSTOM BRAND STYLING (matches the Excel report's dark green) ---
st.markdown("""
<style>
    h1 {
        color: #1B4D3E !important;
        font-family: 'Arial', sans-serif;
        font-weight: 700;
    }
    h3 {
        color: #4A5568 !important;
        font-weight: 400;
    }
    div.stButton > button:first-child {
        background-color: #1B4D3E !important;
        color: white !important;
        border-radius: 6px !important;
        padding: 10px 24px !important;
        font-weight: 600 !important;
        border: none !important;
    }
    div.stButton > button:first-child:hover {
        background-color: #143d31 !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

st.title("HalalQuant")
st.subheader("An Independent Analytical Engine for Quantitative Ratio Screening")

st.warning("""
🚨 **METHODOLOGY & LEGAL DISCLAIMER:** 
This application is an independent analytical simulation that screens quarterly financial 
data based on global mathematical ratio filters.
* **Official Index Compliance:** This software does not directly reflect the official BIST 
Compliance Index or Shariah Board decisions.
* **Sample Variance:** For instance, companies like EREGL and TKNSA are included in the 
official compliance index, but may yield different results here due to data mapping 
differences on third-party sources (Yahoo Finance).
* **Investment Decision:** These results are absolutely not investment advice, binding 
statements, or religious fatwas. Please rely on official stock exchange announcements 
for binding decisions.
""")

@st.cache_data(ttl=1800)
def fetch_financial_data(ticker_symbol):
    try:
        company = ticker_data.Ticker(ticker_symbol)
        return {
            "balance_sheet": company.quarterly_balance_sheet,
            "financials": company.quarterly_financials,
            "history": company.history(period="1y", auto_adjust=False),
            "info": company.info
        }
    except Exception:
        return None


def format_ratio(value, passed, threshold_label):
    """Embed a ✓/✗ symbol alongside the percentage so the pass/fail
    signal doesn't rely on color alone (colorblind accessibility)."""
    if value is None:
        return "– N/A"
    symbol = "✓" if passed else "✗"
    return f"{symbol} %{value:.2f}"


girdi_liste = st.text_input(
    "BIST Ticker Symbols (Separate with commas):", 
    "BIMAS, ASELS, MGROS, THYAO, HEKTS"
)

if st.button("Run Screening"):
    kodlar = [k.strip().upper() for k in girdi_liste.split(",") if k.strip()]
    if not kodlar:
        st.error("Please enter at least one valid ticker symbol.")
    else:
        st.info("⏳ Processing financial metrics over 12-month average prices...")
        toplu_sonuclar = []

        for g in kodlar:
            ticker_symbol = f"{g}.IS" if not g.endswith(".IS") else g
            try:
                data_pack = fetch_financial_data(ticker_symbol)

                if not data_pack:
                    st.warning(f"⚠️ Connection error for {ticker_symbol}. Skipped.")
                    continue

                bilanco_tablosu = data_pack["balance_sheet"]
                gelir_tablosu = data_pack["financials"]
                gecmis_fiyatlar = data_pack["history"]
                shares_outstanding = data_pack["info"].get('sharesOutstanding')

                if (bilanco_tablosu.empty or gelir_tablosu.empty or
                    gecmis_fiyatlar.empty or not shares_outstanding):
                    st.warning(f"⚠️ {ticker_symbol} data is incomplete. Skipped.")
                    continue

                py_f = float(gecmis_fiyatlar['Close'].mean() * shares_outstanding)

                # All four ratios + final verdict now live in screening_logic.py,
                # covered by test_filters.py — see that file before editing the math.
                result = screen_company(bilanco_tablosu, gelir_tablosu, py_f)

                # Most recent quarter this balance sheet actually reports —
                # shown so users know how fresh the underlying data is.
                data_as_of = bilanco_tablosu.columns[0]
                data_as_of_str = data_as_of.strftime("%Y-%m-%d") if hasattr(data_as_of, "strftime") else str(data_as_of)

                toplu_sonuclar.append({
                    "Ticker": g,
                    "Avg Market Cap (1Y)": f"{py_f:,.0f}",
                    "Debt Ratio (%33)": format_ratio(result["debt_ratio"], result["debt_pass"], 33),
                    "Interest Ratio (%5)": format_ratio(result["interest_ratio"], result["interest_pass"], 5),
                    "Cash Ratio (%33)": format_ratio(result["cash_ratio"], result["cash_pass"], 33),
                    "Receivables Ratio (%33)": format_ratio(result["receivables_ratio"], result["receivables_pass"], 33),
                    "FINAL STATUS": result["status"],
                    "Data As Of (Quarter)": data_as_of_str,
                    "_b": result["debt_pass"] if result["debt_ratio"] is not None else None,
                    "_f": result["interest_pass"] if result["interest_ratio"] is not None else None,
                    "_n": result["cash_pass"] if result["cash_ratio"] is not None else None,
                    "_a": result["receivables_pass"] if result["receivables_ratio"] is not None else None,
                })
            except Exception as e:
                st.warning(f"⚠️ Unexpected error while screening {g}: {e}")

        if toplu_sonuclar:
            df_gosterim = pd.DataFrame(toplu_sonuclar).drop(columns=["_b", "_f", "_n", "_a"])
            st.dataframe(df_gosterim, use_container_width=True)
            st.caption(
                "Balance sheet and income statement figures are each company's most recently "
                "reported quarter (see 'Data As Of' column). Market cap is the trailing "
                "12-month average of raw daily closing prices, refreshed every 30 minutes."
            )

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_gosterim.to_excel(writer, index=False, sheet_name='Quarterly Shariah Analysis')
                worksheet = writer.sheets['Quarterly Shariah Analysis']

                n_cols = len(df_gosterim.columns)
                status_col = df_gosterim.columns.get_loc("FINAL STATUS") + 1

                bg_baslik = PatternFill("solid", fgColor="1B4D3E")
                bg_yesil = PatternFill("solid", fgColor="D4EDDA")
                bg_kirmizi = PatternFill("solid", fgColor="F8D7DA")
                bg_gri = PatternFill("solid", fgColor="E2E3E5")

                f_beyaz = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                f_kalin = Font(name="Arial", size=11, bold=True)
                f_normal = Font(name="Arial", size=11)

                for c_idx in range(1, n_cols + 1):
                    cell = worksheet.cell(row=1, column=c_idx)
                    cell.fill = bg_baslik
                    cell.font = f_beyaz
                    cell.alignment = Alignment(horizontal="center")

                ratio_columns = {
                    "Debt Ratio (%33)": "_b",
                    "Interest Ratio (%5)": "_f",
                    "Cash Ratio (%33)": "_n",
                    "Receivables Ratio (%33)": "_a",
                }

                for idx, veri in enumerate(toplu_sonuclar):
                    r_idx = idx + 2
                    for c_idx in range(1, n_cols + 1):
                        worksheet.cell(row=r_idx, column=c_idx).font = f_normal

                    for col_name, flag_key in ratio_columns.items():
                        col_idx = df_gosterim.columns.get_loc(col_name) + 1
                        flag = veri[flag_key]
                        fill = bg_yesil if flag else (bg_kirmizi if flag == False else bg_gri)
                        worksheet.cell(row=r_idx, column=col_idx).fill = fill

                    c_durum = worksheet.cell(row=r_idx, column=status_col)
                    c_durum.font = f_kalin
                    c_durum.fill = (
                        bg_yesil if veri["FINAL STATUS"] == "COMPLIANT"
                        else (bg_kirmizi if veri["FINAL STATUS"] == "NON-COMPLIANT" else bg_gri)
                    )

                for col_idx in range(1, n_cols + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(len(str(cell.value or '')) for cell in worksheet[col_letter])
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 16)

            st.markdown("---")
            st.download_button(
                label="📥 Download Excel Report",
                data=buffer.getvalue(),
                file_name="halalquant_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Data could not be fetched.")
